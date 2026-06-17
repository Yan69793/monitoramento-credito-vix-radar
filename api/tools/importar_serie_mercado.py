#!/usr/bin/env python3
"""
Ingestor de séries históricas de debêntures em múltiplos formatos.

Suporta: Excel Quantum Finance, CSV ANBIMA, CSV B3 UP2DATA.
Normaliza para schema padrão do VIX Radar.

Uso:
    python importar_serie_mercado.py <input_file> [--output OUT.csv] [--upload] \\
        [--endpoint URL] [--admin-senha SENHA] [--dry-run] [--verbose]
"""

import sys
import os
import csv
import json
import argparse
import re
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from typing import List, Dict, Tuple, Optional


def detect_file_format(filepath: str) -> str:
    """Detecta o formato do arquivo de entrada."""
    ext = Path(filepath).suffix.lower()
    if ext == '.xlsx':
        return 'excel'
    elif ext == '.csv':
        with open(filepath, 'r', encoding='utf-8') as f:
            header = f.readline().strip()
            # Heurística: ANBIMA tende a ter campos com PU, Spread, Duration
            # B3 tem padrão similar, mas aqui consideramos CSV genérico
            if 'spread' in header.lower() or 'duration' in header.lower():
                return 'anbima'
            else:
                return 'b3'
    else:
        raise ValueError(f"Formato não reconhecido: {ext}")


def normalize_empresa(nome: str) -> str:
    """Normaliza nome da empresa."""
    return nome.strip().upper()


def normalize_data(data_str: str) -> Optional[str]:
    """Normaliza data para YYYY-MM-DD. Retorna None se inválida."""
    data_str = data_str.strip()
    formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y']
    for fmt in formatos:
        try:
            dt = datetime.strptime(data_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def parse_numero(val: str) -> Optional[float]:
    """Parse de número tolerando vírgula decimal."""
    if not val or not val.strip():
        return None
    val = val.strip().replace(',', '.')
    try:
        return float(val)
    except ValueError:
        return None


def read_excel(filepath: str, verbose: bool = False) -> List[Dict]:
    """Lê arquivo Excel no padrão Quantum Finance."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl não instalado. Execute: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath)
    sheet = None

    # Procura aba "Debêntures" ou variantes
    for name in wb.sheetnames:
        if 'deben' in name.lower():
            sheet = wb[name]
            break
    if not sheet:
        sheet = wb.active

    if verbose:
        print(f"[Excel] Lendo aba '{sheet.title}'")

    rows = []
    header = []

    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i == 1:
            header = [str(c).lower() if c else '' for c in row]
            if verbose:
                print(f"[Excel] Header: {header}")
            continue

        if not any(row):
            continue

        rows.append(dict(zip(header, row)))

    # Mapeamento heurístico de colunas
    data = []
    for row in rows:
        # Detecta colunas com tolerância
        empresa = None
        for col in row:
            if col and any(x in col.lower() for x in ['emiss', 'empresa', 'nome']):
                empresa = row[col]
                break

        data_evt = None
        for col in row:
            if col and any(x in col.lower() for x in ['data', 'data base', 'date']):
                data_evt = row[col]
                break

        spread = None
        for col in row:
            if col and any(x in col.lower() for x in ['spread', 'taxa']):
                spread = parse_numero(str(row[col]) if row[col] else '0')
                if spread and spread < 1:  # Assumir % → bps (×100)
                    spread *= 100
                break

        pu = None
        for col in row:
            if col and any(x in col.lower() for x in ['pu', 'preco']):
                pu = parse_numero(str(row[col]) if row[col] else None)
                break

        duration = None
        for col in row:
            if col and 'duration' in col.lower():
                duration = parse_numero(str(row[col]) if row[col] else None)
                break

        volume = None
        for col in row:
            if col and any(x in col.lower() for x in ['volume', 'valor']):
                volume = parse_numero(str(row[col]) if row[col] else None)
                break

        negocios = None
        for col in row:
            if col and 'negocio' in col.lower():
                negocios = parse_numero(str(row[col]) if row[col] else None)
                break

        maior = None
        for col in row:
            if col and 'maior' in col.lower():
                maior = parse_numero(str(row[col]) if row[col] else None)
                break

        if empresa and data_evt:
            data.append({
                'empresa': empresa,
                'data': data_evt,
                'spread_bps': spread or 0,
                'pu_indicativo': pu or 0,
                'duration': duration or 0,
                'volume': volume or 0,
                'negocios': int(negocios) if negocios else 0,
                'maior_negocio': maior or 0
            })

    return data


def read_csv(filepath: str, verbose: bool = False, filetype: str = 'auto') -> List[Dict]:
    """Lê arquivo CSV nos padrões ANBIMA ou B3."""

    with open(filepath, 'r', encoding='utf-8') as f:
        primeira = f.readline().strip()

    # Detecta separador
    separador = ';' if ';' in primeira else ','
    if verbose:
        print(f"[CSV] Separador detectado: '{separador}'")

    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=separador)
        rows = list(reader)

    if verbose:
        print(f"[CSV] Headers: {list(rows[0].keys()) if rows else 'vazio'}")

    # Mapeamento flexível de colunas (snake_case, camelCase, espaços)
    def find_col(row_keys, patterns):
        for pattern in patterns:
            for key in row_keys:
                if pattern.lower() in key.lower().replace(' ', '').replace('_', ''):
                    return key
        return None

    data = []
    for row in rows:
        if not row or not any(row.values()):
            continue

        keys = list(row.keys())

        empresa_col = find_col(keys, ['emiss', 'empresa', 'nome', 'issuer'])
        data_col = find_col(keys, ['data', 'referencia', 'base', 'date'])
        spread_col = find_col(keys, ['spread', 'taxa', 'rate'])
        pu_col = find_col(keys, ['pu', 'preco', 'price'])
        duration_col = find_col(keys, ['duration', 'duracao', 'macaulay'])
        volume_col = find_col(keys, ['volume', 'valor', 'amount'])
        negocios_col = find_col(keys, ['negocio', 'trades', 'qty'])
        maior_col = find_col(keys, ['maior', 'max'])

        empresa = row.get(empresa_col, '').strip() if empresa_col else None
        data_evt = row.get(data_col, '').strip() if data_col else None

        if not empresa or not data_evt:
            continue

        spread = parse_numero(row.get(spread_col, '0') if spread_col else '0')
        if spread and spread < 1:
            spread *= 100

        data.append({
            'empresa': empresa,
            'data': data_evt,
            'spread_bps': spread or 0,
            'pu_indicativo': parse_numero(row.get(pu_col, '0') if pu_col else '0') or 0,
            'duration': parse_numero(row.get(duration_col, '0') if duration_col else '0') or 0,
            'volume': parse_numero(row.get(volume_col, '0') if volume_col else '0') or 0,
            'negocios': int(parse_numero(row.get(negocios_col, '0') if negocios_col else '0') or 0),
            'maior_negocio': parse_numero(row.get(maior_col, '0') if maior_col else '0') or 0
        })

    return data


def normalize_records(raw_records: List[Dict], verbose: bool = False) -> Tuple[List[Dict], List[str]]:
    """Normaliza registros brutos: valida datas, deduplica, consolida."""

    normalized = []
    rejeitos = []
    visto = {}  # (empresa, data) -> idx do último

    for rec in raw_records:
        empresa = normalize_empresa(rec.get('empresa', ''))
        data = normalize_data(str(rec.get('data', '')))

        if not empresa:
            rejeitos.append(f"Empresa vazia em {rec}")
            continue

        if not data:
            rejeitos.append(f"Data inválida para {empresa}: {rec.get('data')}")
            continue

        key = (empresa, data)
        if key in visto:
            # Sobrescreve anterior (última ocorrência vence)
            idx = visto[key]
            normalized[idx] = {
                'empresa': empresa,
                'data': data,
                'spread_bps': rec.get('spread_bps', 0),
                'pu_indicativo': rec.get('pu_indicativo', 0),
                'duration': rec.get('duration', 0),
                'volume': rec.get('volume', 0),
                'negocios': rec.get('negocios', 0),
                'maior_negocio': rec.get('maior_negocio', 0)
            }
        else:
            visto[key] = len(normalized)
            normalized.append({
                'empresa': empresa,
                'data': data,
                'spread_bps': rec.get('spread_bps', 0),
                'pu_indicativo': rec.get('pu_indicativo', 0),
                'duration': rec.get('duration', 0),
                'volume': rec.get('volume', 0),
                'negocios': rec.get('negocios', 0),
                'maior_negocio': rec.get('maior_negocio', 0)
            })

    if verbose and rejeitos:
        print(f"[Validação] {len(rejeitos)} registros rejeitados")
        for r in rejeitos[:5]:
            print(f"  - {r}")

    return normalized, rejeitos


def export_csv(records: List[Dict], output_path: str) -> None:
    """Exporta para CSV no schema padrão."""

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(
            f,
            fieldnames=['empresa', 'data', 'spread_bps', 'pu_indicativo', 'duration', 'volume', 'negocios', 'maior_negocio'],
            delimiter=';'
        )
        writer.writeheader()
        for rec in sorted(records, key=lambda r: (r['empresa'], r['data'])):
            writer.writerow(rec)


def upload_csv(csv_path: str, endpoint: str, admin_senha: str, verbose: bool = False, dry_run: bool = False) -> None:
    """Faz upload do CSV ao Worker via endpoint admin."""

    try:
        import requests
    except ImportError:
        raise RuntimeError("requests não instalado. Execute: pip install requests")

    with open(csv_path, 'r', encoding='utf-8') as f:
        csv_content = f.read()

    # Quebra em batches de ~50 linhas
    lines = csv_content.split('\n')
    header = lines[0]
    data_lines = [l for l in lines[1:] if l.strip()]

    batch_size = 50
    batches = []
    for i in range(0, len(data_lines), batch_size):
        batch_lines = data_lines[i:i+batch_size]
        batch_csv = header + '\n' + '\n'.join(batch_lines)
        batches.append(batch_csv)

    if verbose:
        print(f"[Upload] {len(batches)} batch(es) para enviar")

    if dry_run:
        print("[DRY-RUN] Não enviando dados")
        return

    for idx, batch_csv in enumerate(batches, 1):
        payload = {
            'action': 'admin_ingestao_csv',
            'admin_senha': admin_senha,
            'csv': batch_csv
        }

        retry = 0
        while retry < 3:
            try:
                if verbose:
                    print(f"[Upload] Batch {idx}/{len(batches)} (tentativa {retry+1}/3)")

                resp = requests.post(endpoint, json=payload, timeout=30)

                if resp.status_code == 404:
                    raise RuntimeError(
                        f"Endpoint {endpoint} retornou 404. "
                        "O Worker ainda não foi deployado com v4.8.8. "
                        "Aplique o patch antes de usar --upload."
                    )

                if resp.status_code >= 400:
                    print(f"[Erro] HTTP {resp.status_code}: {resp.text[:200]}")
                    retry += 1
                    continue

                result = resp.json()
                if verbose:
                    print(f"[OK] Batch {idx}: {result.get('message', 'sucesso')}")
                break

            except requests.RequestException as e:
                retry += 1
                if retry < 3:
                    wait = 2 ** retry
                    if verbose:
                        print(f"[Erro] {e}. Aguardando {wait}s...")
                    import time
                    time.sleep(wait)
                else:
                    raise RuntimeError(f"Upload falhou após 3 tentativas: {e}")


def main():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('input_file', help='Arquivo de entrada (.xlsx ou .csv)')
    parser.add_argument('--output', default='serie_mercado_normalizado.csv', help='Caminho de saída CSV')
    parser.add_argument('--upload', action='store_true', help='Faz upload ao Worker após normalizar')
    parser.add_argument('--endpoint', default='https://radar-credito-api.prospects-intel.workers.dev', help='Endpoint admin do Worker')
    parser.add_argument('--admin-senha', default='', help='Senha admin para upload')
    parser.add_argument('--dry-run', action='store_true', help='Não escreve/envia, apenas simula')
    parser.add_argument('--verbose', action='store_true', help='Saída detalhada')

    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"Erro: arquivo '{args.input_file}' não encontrado", file=sys.stderr)
        sys.exit(1)

    print(f"[Ingestão] Lendo {args.input_file}")

    # Detecta e lê
    file_type = detect_file_format(args.input_file)
    print(f"[Formato] {file_type.upper()}")

    if file_type == 'excel':
        raw = read_excel(args.input_file, verbose=args.verbose)
    else:
        raw = read_csv(args.input_file, verbose=args.verbose, filetype=file_type)

    print(f"[Raw] {len(raw)} registros lidos")

    # Normaliza
    normalized, rejeitos = normalize_records(raw, verbose=args.verbose)
    print(f"[Normalizado] {len(normalized)} registros válidos, {len(rejeitos)} rejeitados")

    # Calcula período e estatísticas
    if normalized:
        datas = [r['data'] for r in normalized]
        periodo = f"{min(datas)} a {max(datas)}"
        empresas = set(r['empresa'] for r in normalized)
        print(f"[Período] {periodo}")
        print(f"[Empresas] {len(empresas)} únicas: {', '.join(sorted(empresas)[:5])}{'...' if len(empresas) > 5 else ''}")

    # Exporta
    if args.dry_run:
        print("[DRY-RUN] Não escrevendo CSV")
    else:
        export_csv(normalized, args.output)
        print(f"[Saída] {args.output} ({len(normalized)} linhas)")

    # Upload (se solicitado)
    if args.upload:
        if not args.admin_senha:
            print("Erro: --upload exige --admin-senha", file=sys.stderr)
            sys.exit(1)

        upload_csv(args.output if not args.dry_run else args.output, args.endpoint, args.admin_senha, verbose=args.verbose, dry_run=args.dry_run)

    print("[Concluído]")


if __name__ == '__main__':
    main()
